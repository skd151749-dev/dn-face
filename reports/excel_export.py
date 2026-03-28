"""
DN FACE - Report Export Module
Creates Excel, CSV, and lightweight PDF attendance reports.
"""

import os
from datetime import datetime
from typing import Dict, List, Optional

import pandas as pd

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "reports", "exports")

DISPLAY_COLUMNS = [
    "Name",
    "ID",
    "Group",
    "Sex",
    "Role",
    "Class/Dept",
    "Date",
    "Morning Check-in",
    "Morning Status",
    "Morning Check-out",
    "Afternoon Check-in",
    "Afternoon Status",
    "Afternoon Check-out",
    "Overall Status",
    "Early Leave",
]

COLUMN_MAP = {
    "name": "Name",
    "user_id": "ID",
    "group_name": "Group",
    "sex": "Sex",
    "role": "Role",
    "class_dept": "Class/Dept",
    "date": "Date",
    "morning_check_in": "Morning Check-in",
    "morning_status": "Morning Status",
    "morning_check_out": "Morning Check-out",
    "afternoon_check_in": "Afternoon Check-in",
    "afternoon_status": "Afternoon Status",
    "afternoon_check_out": "Afternoon Check-out",
    "late_status": "Overall Status",
    "early_leave": "Early Leave",
}

PDF_COLUMNS = [
    ("Date", 10),
    ("ID", 10),
    ("Name", 18),
    ("Group", 12),
    ("AM In", 8),
    ("AM Status", 8),
    ("AM Out", 8),
    ("PM In", 8),
    ("PM Status", 8),
    ("PM Out", 8),
    ("Overall", 8),
    ("Early Leave", 8),
]


def _slug(value: Optional[str], fallback: str = "all-groups") -> str:
    raw = (value or fallback).strip().lower()
    chars = [ch if ch.isalnum() else "-" for ch in raw]
    collapsed = "".join(chars).strip("-")
    return collapsed or fallback


def _pdf_escape(text: str) -> str:
    return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _fit(value, width: int) -> str:
    text = str(value or "-")
    if len(text) <= width:
        return text.ljust(width)
    if width <= 3:
        return text[:width]
    return f"{text[:width - 3]}..."


def _build_pdf(pages: List[List[str]]) -> bytes:
    objects: List[bytes] = []

    def add_object(payload: bytes) -> int:
        objects.append(payload)
        return len(objects)

    font_id = add_object(b"<< /Type /Font /Subtype /Type1 /BaseFont /Courier >>")
    page_refs = []

    for page_lines in pages:
        content = "BT\n/F1 10 Tf\n36 806 Td\n13 TL\n"
        for line in page_lines:
            content += f"({_pdf_escape(line)}) Tj\nT*\n"
        content += "ET"
        content_bytes = content.encode("latin-1", "replace")
        content_id = add_object(
            f"<< /Length {len(content_bytes)} >>\nstream\n".encode("latin-1")
            + content_bytes
            + b"\nendstream"
        )
        page_id = add_object(
            f"<< /Type /Page /Parent {{PAGES}} 0 R /MediaBox [0 0 612 842] "
            f"/Resources << /Font << /F1 {font_id} 0 R >> >> /Contents {content_id} 0 R >>".encode("latin-1")
        )
        page_refs.append(page_id)

    kids = " ".join(f"{page_id} 0 R" for page_id in page_refs)
    pages_id = add_object(f"<< /Type /Pages /Kids [{kids}] /Count {len(page_refs)} >>".encode("latin-1"))
    catalog_id = add_object(f"<< /Type /Catalog /Pages {pages_id} 0 R >>".encode("latin-1"))

    final_objects = []
    for payload in objects:
        if b"{PAGES}" in payload:
            payload = payload.replace(b"{PAGES}", str(pages_id).encode("latin-1"))
        final_objects.append(payload)

    buffer = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for index, payload in enumerate(final_objects, start=1):
        offsets.append(len(buffer))
        buffer.extend(f"{index} 0 obj\n".encode("latin-1"))
        buffer.extend(payload)
        buffer.extend(b"\nendobj\n")

    xref_offset = len(buffer)
    buffer.extend(f"xref\n0 {len(final_objects) + 1}\n".encode("latin-1"))
    buffer.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        buffer.extend(f"{offset:010d} 00000 n \n".encode("latin-1"))
    buffer.extend(
        f"trailer\n<< /Size {len(final_objects) + 1} /Root {catalog_id} 0 R >>\nstartxref\n{xref_offset}\n%%EOF".encode(
            "latin-1"
        )
    )
    return bytes(buffer)


class ExcelExporter:
    def __init__(self):
        os.makedirs(OUTPUT_DIR, exist_ok=True)

    def export(
        self,
        records: List[Dict],
        period: str = "daily",
        format: str = "excel",
        group: Optional[str] = None,
        summary: Optional[Dict] = None,
    ) -> str:
        normalized = self._normalize_dataframe(records)
        report_format = (format or "excel").strip().lower()
        if report_format in {"xlsx", "excel"}:
            return self._export_excel(normalized, period, group, summary)
        if report_format == "csv":
            return self._export_csv(normalized, period, group)
        if report_format == "pdf":
            return self._export_pdf(normalized, period, group, summary)
        raise ValueError("Unsupported report format")

    def _normalize_dataframe(self, records: List[Dict]) -> pd.DataFrame:
        if not records:
            df = pd.DataFrame(columns=DISPLAY_COLUMNS)
        else:
            df = pd.DataFrame(records).rename(columns=COLUMN_MAP)
            for column in DISPLAY_COLUMNS:
                if column not in df.columns:
                    df[column] = ""
            df = df[DISPLAY_COLUMNS]
            df = df.fillna("")
            for text_col in [
                "Morning Check-in",
                "Morning Status",
                "Morning Check-out",
                "Afternoon Check-in",
                "Afternoon Status",
                "Afternoon Check-out",
                "Overall Status",
                "Early Leave",
                "Group",
                "Class/Dept",
                "Sex",
            ]:
                df[text_col] = df[text_col].replace("", "-")
        return df

    def _build_filepath(self, period: str, extension: str, group: Optional[str] = None) -> str:
        filename = f"attendance_{period}_{_slug(group)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{extension}"
        return os.path.abspath(os.path.join(OUTPUT_DIR, filename))

    def _report_title(self, period: str, group: Optional[str], summary: Optional[Dict]) -> str:
        group_label = group or "All Groups"
        range_info = (summary or {}).get("range") or {}
        start = range_info.get("start")
        end = range_info.get("end")
        if start and end:
            return f"DN FACE Attendance Report - {period.capitalize()} - {group_label} ({start} to {end})"
        return f"DN FACE Attendance Report - {period.capitalize()} - {group_label}"

    def _export_excel(self, df: pd.DataFrame, period: str, group: Optional[str], summary: Optional[Dict]) -> str:
        filepath = self._build_filepath(period, "xlsx", group)
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Attendance", index=False, startrow=2)
            worksheet = writer.sheets["Attendance"]

            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter

            title = self._report_title(period, group, summary)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(df.columns)))
            title_cell = worksheet.cell(row=1, column=1)
            title_cell.value = title
            title_cell.font = Font(color="FFFFFF", bold=True, size=14)
            title_cell.fill = PatternFill(start_color="0B132B", end_color="0B132B", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            worksheet.row_dimensions[1].height = 28

            if summary:
                summary_cell = worksheet.cell(row=2, column=1)
                summary_cell.value = (
                    f"Total Users: {summary.get('total_users', 0)} | "
                    f"Present: {summary.get('present', 0)} | "
                    f"Late: {summary.get('late', 0)} | "
                    f"Absent: {summary.get('absent', 0)}"
                )
                summary_cell.font = Font(color="A5F3FC", italic=True, size=10)

            header_fill = PatternFill(start_color="172554", end_color="172554", fill_type="solid")
            header_font = Font(color="CFFAFE", bold=True, size=11)
            for col_num, column in enumerate(df.columns, 1):
                cell = worksheet.cell(row=3, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            fill_even = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
            fill_odd = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
            for row_num in range(4, len(df) + 4):
                fill = fill_even if row_num % 2 == 0 else fill_odd
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.fill = fill
                    cell.font = Font(color="E5E7EB", size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            for col_num, column in enumerate(df.columns, 1):
                max_len = max(len(str(column)), df[column].astype(str).map(len).max() if len(df) else 0)
                worksheet.column_dimensions[get_column_letter(col_num)].width = min(max_len + 4, 28)

            worksheet.freeze_panes = "A4"

        return filepath

    def _export_csv(self, df: pd.DataFrame, period: str, group: Optional[str]) -> str:
        filepath = self._build_filepath(period, "csv", group)
        df.to_csv(filepath, index=False)
        return filepath

    def _export_pdf(self, df: pd.DataFrame, period: str, group: Optional[str], summary: Optional[Dict]) -> str:
        filepath = self._build_filepath(period, "pdf", group)
        title = self._report_title(period, group, summary)
        summary_line = (
            f"Total Users: {summary.get('total_users', 0)} | Present: {summary.get('present', 0)} | "
            f"Late: {summary.get('late', 0)} | Absent: {summary.get('absent', 0)}"
            if summary
            else "No summary available"
        )

        pdf_df = df.rename(
            columns={
                "Morning Check-in": "AM In",
                "Morning Status": "AM Status",
                "Morning Check-out": "AM Out",
                "Afternoon Check-in": "PM In",
                "Afternoon Status": "PM Status",
                "Afternoon Check-out": "PM Out",
                "Overall Status": "Overall",
            }
        )

        header = " | ".join(_fit(label, width) for label, width in PDF_COLUMNS)
        divider = "-" * len(header)

        lines = [title, summary_line, "", header, divider]
        for _, row in pdf_df.iterrows():
            line = " | ".join(_fit(row[label], width) for label, width in PDF_COLUMNS)
            lines.append(line)

        if len(lines) == 5:
            lines.append("No records found for the selected period and group.")

        lines_per_page = 45
        pages = [lines[index:index + lines_per_page] for index in range(0, len(lines), lines_per_page)]
        pdf_bytes = _build_pdf(pages)
        with open(filepath, "wb") as handle:
            handle.write(pdf_bytes)
        return filepath
